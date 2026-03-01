#!/usr/bin/env python3
"""
generate_excel_output.py - Generate Excel file with keeper recommendations

Creates a formatted Excel workbook with:
1. The Nudes keeper recommendations
2. All player valuations by position
3. Position scarcity analysis
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# League rules
MAX_YEARS_KEPT = 3
INELIGIBLE_ROUNDS = [1, 2, 3]
UNDRAFTED_ROUND = 18
NUM_KEEPERS = 8

# Draft round value curve
DRAFT_ROUND_VALUES = {
    1: 50, 2: 42, 3: 35, 4: 30, 5: 26, 6: 23, 7: 20, 8: 18,
    9: 16, 10: 14, 11: 12, 12: 10, 13: 9, 14: 8, 15: 7,
    16: 6, 17: 5, 18: 4, 19: 3, 20: 2, 21: 2, 22: 1, 23: 1, 24: 1, 25: 1
}


def calculate_keeper_cost(draft_round: int) -> int:
    if draft_round == 0:
        draft_round = UNDRAFTED_ROUND
    return max(1, draft_round - 3)


def get_round_value(round_num: int) -> float:
    round_num = max(1, min(25, round_num))
    return DRAFT_ROUND_VALUES.get(round_num, 1)


def generate_excel():
    data_dir = Path(__file__).parent / 'data'
    proj_dir = data_dir / 'projections'
    roster_dir = data_dir / 'rosters'
    output_dir = Path(__file__).parent / 'output'
    output_dir.mkdir(exist_ok=True)

    # Load data
    valuations = pd.read_csv(proj_dir / 'sgp_player_values_v3.csv')
    roster = pd.read_csv(roster_dir / 'yahoo_league.csv')

    # Filter to The Nudes
    my_roster = roster[roster['Team'] == 'The Nudes'].copy()

    # Create workbook
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=12, color='FFFFFF')
    keeper_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===== Sheet 1: Keeper Recommendations =====
    ws = wb.active
    ws.title = "Keeper Recommendations"

    # Build keeper analysis
    analysis = []
    ineligible = []

    for _, row in my_roster.iterrows():
        player = row['Player']
        position = row['Position']
        draft_round = int(row['DraftRound'])
        years_kept = int(row['YearsKept'])

        # Check eligibility
        is_ineligible = False
        reason = ""

        if draft_round in INELIGIBLE_ROUNDS:
            is_ineligible = True
            reason = f"Rounds 1-3 ineligible"
        elif years_kept >= MAX_YEARS_KEPT:
            is_ineligible = True
            reason = f"3 years already kept"

        # Find player value
        match = valuations[valuations['Name'].str.lower() == player.lower()]
        if len(match) == 0:
            # Try fuzzy
            from difflib import SequenceMatcher
            best_score = 0
            for _, v in valuations.iterrows():
                score = SequenceMatcher(None, player.lower(), v['Name'].lower()).ratio()
                if score > best_score and score > 0.8:
                    best_score = score
                    match = valuations[valuations['Name'] == v['Name']]

        if len(match) == 0:
            ineligible.append({
                'Player': player, 'Position': position,
                'Reason': 'Not in projections', 'Value': None
            })
            continue

        value_row = match.iloc[0]
        value = value_row['dollar_value']
        rank = int(value_row['overall_rank'])

        if is_ineligible:
            ineligible.append({
                'Player': player, 'Position': position,
                'Reason': reason, 'Value': value, 'Rank': rank
            })
            continue

        keeper_round = calculate_keeper_cost(draft_round if draft_round > 0 else UNDRAFTED_ROUND)
        keeper_cost = get_round_value(keeper_round)
        surplus = value - keeper_cost
        years_remaining = MAX_YEARS_KEPT - years_kept

        analysis.append({
            'Player': player,
            'Position': position,
            'Value': value,
            'Rank': rank,
            'Draft Round': draft_round if draft_round > 0 else 'UD',
            'Keeper Round': keeper_round,
            'Keeper Cost': keeper_cost,
            'Surplus': surplus,
            'Years Kept': years_kept,
            'Years Left': years_remaining
        })

    # Sort by surplus
    analysis_df = pd.DataFrame(analysis)
    analysis_df = analysis_df.sort_values('Surplus', ascending=False)

    # Write header
    ws['A1'] = "THE NUDES - KEEPER RECOMMENDATIONS"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:J1')

    ws['A3'] = f"Top {NUM_KEEPERS} Recommended Keepers (sorted by surplus value)"
    ws['A3'].font = Font(bold=True, size=12)

    # Headers
    headers = ['#', 'Player', 'Position', 'Value', 'Rank', 'Draft Rd', 'Keep Rd', 'Cost', 'Surplus', 'Yrs Left']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for i, (_, row) in enumerate(analysis_df.iterrows(), 1):
        row_num = 5 + i
        is_keeper = i <= NUM_KEEPERS

        values = [
            i,
            row['Player'],
            row['Position'],
            f"${row['Value']:.1f}",
            row['Rank'],
            row['Draft Round'],
            row['Keeper Round'],
            f"${row['Keeper Cost']:.0f}",
            f"${row['Surplus']:.1f}",
            row['Years Left']
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = thin_border
            if is_keeper:
                cell.fill = keeper_fill

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 10

    # Ineligible section
    inelig_start = 5 + len(analysis_df) + 3
    ws.cell(row=inelig_start, column=1, value="Ineligible Players").font = Font(bold=True, size=12)

    for i, p in enumerate(ineligible, 1):
        row_num = inelig_start + i
        ws.cell(row=row_num, column=1, value=p['Player'])
        ws.cell(row=row_num, column=2, value=p['Position'])
        ws.cell(row=row_num, column=3, value=f"${p['Value']:.1f}" if p['Value'] else "N/A")
        ws.cell(row=row_num, column=4, value=p['Reason'])

    # ===== Sheet 2: All Player Values =====
    ws2 = wb.create_sheet("All Players")

    headers2 = ['Rank', 'Player', 'Type', 'Position', 'Value', 'SGP', 'Multiplier']
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for i, (_, row) in enumerate(valuations.head(200).iterrows(), 2):
        ws2.cell(row=i, column=1, value=int(row['overall_rank']))
        ws2.cell(row=i, column=2, value=row['Name'])
        ws2.cell(row=i, column=3, value=row['player_type'])
        ws2.cell(row=i, column=4, value=row.get('primary_position', ''))
        ws2.cell(row=i, column=5, value=f"${row['dollar_value']:.1f}")
        ws2.cell(row=i, column=6, value=f"{row['total_sgp']:.1f}")
        ws2.cell(row=i, column=7, value=f"{row.get('position_multiplier', 1.0):.2f}")

    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 10
    ws2.column_dimensions['D'].width = 10
    ws2.column_dimensions['E'].width = 10
    ws2.column_dimensions['F'].width = 8
    ws2.column_dimensions['G'].width = 10

    # ===== Sheet 3: Catchers =====
    ws3 = wb.create_sheet("Catchers")
    catchers = valuations[valuations['primary_position'] == 'C'].head(20)

    ws3['A1'] = "Top Catchers (1.35x scarcity multiplier)"
    ws3['A1'].font = Font(bold=True, size=14)

    headers3 = ['Rank', 'Player', 'Value', 'Base SGP', 'Adjusted SGP']
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill

    for i, (_, row) in enumerate(catchers.iterrows(), 4):
        ws3.cell(row=i, column=1, value=int(row['overall_rank']))
        ws3.cell(row=i, column=2, value=row['Name'])
        ws3.cell(row=i, column=3, value=f"${row['dollar_value']:.1f}")
        ws3.cell(row=i, column=4, value=f"{row['total_sgp']:.1f}")
        ws3.cell(row=i, column=5, value=f"{row.get('adjusted_sgp', row['total_sgp']):.1f}")

    ws3.column_dimensions['B'].width = 25

    # Save
    output_file = output_dir / 'keeper_recommendations.xlsx'
    wb.save(output_file)
    print(f"\nExcel file saved to: {output_file}")

    # Print summary
    print("\nContents:")
    print("  1. Keeper Recommendations - Your top keepers sorted by surplus value")
    print("  2. All Players - Top 200 players by value")
    print("  3. Catchers - Top catchers with scarcity adjustment")


if __name__ == '__main__':
    generate_excel()
